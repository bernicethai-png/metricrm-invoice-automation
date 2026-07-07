#!/usr/bin/env python3
"""
MetriCRM Invoice Generator - 从Excel模板直接生成PDF发票
"""

import os
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ===== 配置 =====
TEMPLATE_PATH = Path(__file__).parent / "MetriCRM.xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_PDF = OUTPUT_DIR / "output_invoice.pdf"

def get_malaysia_date():
    """获取马来西亚当前日期"""
    utc_now = datetime.utcnow()
    malaysia_now = utc_now + timedelta(hours=8)
    return malaysia_now.date()

def get_invoice_info():
    """获取发票信息"""
    today = get_malaysia_date()
    year = today.year
    month = today.month
    day = today.day

    invoice_num = f"WB_INV_{year%100:02d}{month:02d}{day:02d}_02"

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    invoice_date = f"{day:02d}-{months[month-1]}-{year%100:02d}"

    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    billing_period = f"01 {months[month-1]} {year} To {last_day} {months[month-1]} {year}"

    return {
        "invoice_num": invoice_num,
        "invoice_date": invoice_date,
        "billing_period": billing_period,
        "year": year,
        "month": month,
        "day": day
    }

def read_excel_data():
    """从Excel读取发票数据"""
    if not TEMPLATE_PATH.exists():
        return None

    try:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # 读取基本信息
        data = {
            "company_name": ws['C2'].value,
            "address1": ws['C3'].value,
            "address2": ws['C4'].value,
            "postal": ws['C5'].value,
            "email": ws['C6'].value,
            "phone": ws['C7'].value,
            "bill_to_name": ws['B10'].value,
            "bill_to_addr1": ws['B11'].value,
            "bill_to_addr2": ws['B12'].value,
            "items": []
        }

        # 读取项目（行19-20）
        for row in range(19, 21):
            desc = ws[f'B{row}'].value
            qty = ws[f'F{row}'].value
            price = ws[f'G{row}'].value
            if desc and qty and price:
                data["items"].append({
                    "desc": desc,
                    "qty": qty,
                    "price": price,
                    "total": qty * price
                })

        return data
    except Exception as e:
        print(f"❌ 读取Excel失败: {e}")
        return None

def generate_pdf():
    """生成PDF发票"""
    OUTPUT_DIR.mkdir(exist_ok=True)

    print("🚀 开始生成发票...")

    # 获取发票信息
    invoice_info = get_invoice_info()
    print(f"📋 发票信息:")
    print(f"   号码: {invoice_info['invoice_num']}")
    print(f"   日期: {invoice_info['invoice_date']}")
    print(f"   周期: {invoice_info['billing_period']}")

    # 读取Excel数据
    print("📂 加载Excel模板...")
    data = read_excel_data()
    if not data:
        print("❌ 无法读取Excel文件")
        return False

    try:
        # 创建PDF文档
        doc = SimpleDocTemplate(str(OUTPUT_PDF), pagesize=letter,
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)

        story = []
        styles = getSampleStyleSheet()

        # 自定义样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1a8917'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # 公司信息和发票标题
        header_data = [
            [Paragraph(f"<b>{data['company_name']}</b>", styles['Normal']),
             None, None,
             Paragraph("<b style='font-size:18px'>INVOICE</b>", title_style)],
            [Paragraph(str(data['address1']), styles['Normal']),
             None, None,
             Paragraph(f"<b>DATE</b>: {invoice_info['invoice_date']}", styles['Normal'])],
            [Paragraph(f"{data['address2']}, {data['postal']}", styles['Normal']),
             None, None,
             Paragraph(f"<b>INVOICE #</b>: {invoice_info['invoice_num']}", styles['Normal'])],
            [Paragraph(f"EMAIL: {data['email']}", styles['Normal']),
             None, None,
             Paragraph(f"<b>TERMS</b>: 30 days", styles['Normal'])],
            [Paragraph(f"PHONE: {data['phone']}", styles['Normal']),
             None, None, None],
        ]

        header_table = Table(header_data, colWidths=[2.5*inch, 0.5*inch, 0.5*inch, 2*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 11),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        story.append(header_table)
        story.append(Spacer(1, 0.2*inch))

        # 账单信息
        bill_data = [
            [Paragraph("<b>BILL TO:</b>", styles['Normal']), None],
            [Paragraph(str(data['bill_to_name']), styles['Normal']), None],
            [Paragraph(str(data['bill_to_addr1']), styles['Normal']), None],
            [Paragraph(str(data['bill_to_addr2']), styles['Normal']), None],
        ]

        bill_table = Table(bill_data, colWidths=[4*inch, 2.5*inch])
        story.append(bill_table)
        story.append(Spacer(1, 0.2*inch))

        # 发票项目表
        items_data = [
            [Paragraph("<b>DESCRIPTION</b>", styles['Normal']),
             Paragraph("<b>QTY</b>", styles['Normal']),
             Paragraph("<b>Price</b>", styles['Normal']),
             Paragraph("<b>TOTAL AMOUNT(RM)</b>", styles['Normal'])]
        ]

        for item in data['items']:
            items_data.append([
                Paragraph(str(item['desc']), styles['Normal']),
                Paragraph(str(item['qty']), styles['Normal']),
                Paragraph(f"{item['price']:.2f}", styles['Normal']),
                Paragraph(f"{item['total']:.2f}", styles['Normal'])
            ])

        # 添加计费周期
        items_data.append([
            Paragraph(f"<i>{invoice_info['billing_period']}</i>", styles['Normal']),
            None, None, None
        ])

        items_table = Table(items_data, colWidths=[3*inch, 0.7*inch, 0.8*inch, 1.5*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#cccccc')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -2), 1, colors.black),
            ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
            ('ALIGN', (1, -2), (-1, -2), 'RIGHT'),
            ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
        ]))

        story.append(items_table)
        story.append(Spacer(1, 0.2*inch))

        # 总计
        total = sum(item['total'] for item in data['items'])
        total_data = [
            [Paragraph("<b>TOTAL(RM)</b>", styles['Normal']),
             None, None,
             Paragraph(f"<b>{total:.2f}</b>", styles['Normal'])]
        ]

        total_table = Table(total_data, colWidths=[3*inch, 0.7*inch, 0.8*inch, 1.5*inch])
        total_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ]))

        story.append(total_table)
        story.append(Spacer(1, 0.3*inch))

        # 银行信息
        bank_info = [
            [Paragraph("<b>Bank Detail</b>", styles['Normal'])],
            [Paragraph("Webbalances solution", styles['Normal'])],
            [Paragraph("8010 438 379", styles['Normal'])],
            [Paragraph("CIMB BANK", styles['Normal'])],
        ]

        bank_table = Table(bank_info, colWidths=[2*inch])
        bank_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))

        story.append(bank_table)
        story.append(Spacer(1, 0.2*inch))

        # 底部声明
        story.append(Paragraph("<i>This invoice is generated by computer. No signature needed</i>",
                             ParagraphStyle('Footer', parent=styles['Normal'],
                                          fontSize=8, alignment=TA_CENTER)))

        # 生成PDF
        print("📝 生成PDF...")
        doc.build(story)

        print(f"✅ PDF生成成功: {OUTPUT_PDF}")
        return True

    except Exception as e:
        print(f"❌ 生成失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = generate_pdf()
    exit(0 if success else 1)
