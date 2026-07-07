#!/usr/bin/env python3
"""
MetriCRM Invoice Generator - 从Excel模板生成PDF
使用LibreOffice在后台转换Excel为PDF
"""

import os
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook

# ===== 配置 =====
TEMPLATE_PATH = Path(__file__).parent / "MetriCRM.xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"
TEMP_EXCEL = OUTPUT_DIR / "temp_invoice.xlsx"
OUTPUT_PDF = OUTPUT_DIR / "output_invoice.pdf"

def get_malaysia_date():
    """获取马来西亚当前日期"""
    utc_now = datetime.utcnow()
    malaysia_now = utc_now + timedelta(hours=8)
    return malaysia_now.date()

def get_invoice_info():
    """获取发票信息"""
    today = get_malaysia_date()
    year = today.year
    month = today.month
    day = today.day

    invoice_num = f"WB_INV_{year%100:02d}{month:02d}{day:02d}_02"

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    invoice_date = f"{day:02d}-{months[month-1]}-{year%100:02d}"

    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    billing_period = f"01 {months[month-1]} {year} To {last_day} {months[month-1]} {year}"

    return {
        "invoice_num": invoice_num,
        "invoice_date": invoice_date,
        "billing_period": billing_period,
    }

def generate_pdf():
    """生成PDF发票"""
    OUTPUT_DIR.mkdir(exist_ok=True)

    print("🚀 开始生成发票...")

    if not TEMPLATE_PATH.exists():
        print(f"❌ 模板文件不存在: {TEMPLATE_PATH}")
        return False

    invoice_info = get_invoice_info()
    print(f"📋 发票信息:")
    print(f"   号码: {invoice_info['invoice_num']}")
    print(f"   日期: {invoice_info['invoice_date']}")
    print(f"   周期: {invoice_info['billing_period']}")

    try:
        # 加载模板
        print("📂 加载Excel模板...")
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # 修改发票信息
        print("✏️  修改发票字段...")
        ws['H10'] = invoice_info['invoice_date']
        ws['H11'] = invoice_info['invoice_num']
        ws['B21'] = invoice_info['billing_period']

        # 保存临时Excel
        print(f"💾 保存临时Excel: {TEMP_EXCEL}")
        wb.save(TEMP_EXCEL)

        # 用LibreOffice转换为PDF
        print("🔄 转换Excel为PDF...")

        # 方法1: 用libreoffice命令行
        result = subprocess.run(
            [
                'libreoffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(OUTPUT_DIR),
                str(TEMP_EXCEL)
            ],
            capture_output=True,
            text=True,
            timeout=60,
            env={**os.environ, 'DISPLAY': ':99', 'HOME': '/tmp'}
        )

        print(f"Return code: {result.returncode}")
        if result.stdout:
            print(f"STDOUT: {result.stdout}")
        if result.stderr:
            print(f"STDERR: {result.stderr}")

        # 检查是否生成了PDF
        if not OUTPUT_PDF.exists():
            # 尝试找其他可能的输出文件
            possible_pdf = OUTPUT_DIR / "temp_invoice.pdf"
            if possible_pdf.exists():
                possible_pdf.rename(OUTPUT_PDF)
                print(f"✅ PDF生成成功: {OUTPUT_PDF}")
            else:
                print(f"❌ PDF生成失败: 找不到输出文件")
                return False
        else:
            print(f"✅ PDF生成成功: {OUTPUT_PDF}")

        # 清理临时文件
        try:
            TEMP_EXCEL.unlink()
        except:
            pass

        return True

    except Exception as e:
        print(f"❌ 生成失败: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = generate_pdf()
    exit(0 if success else 1)
